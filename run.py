from typing import List
import chromedriver_autoinstaller
import numpy as np
import pandas as pd
import requests
import time
import matplotlib.pyplot as plt
from openpyxl import load_workbook, Workbook, utils
from io import BytesIO
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

#%%
# 한글 폰트 오류 해결

from matplotlib import font_manager, rc
font_path = "./malgun.ttf"   #폰트파일의 위치
font_name = font_manager.FontProperties(fname=font_path).get_name()
rc('font', family=font_name)

#%%
def get_excel():
    '''selenium으로 웹에서 excel 가져오기'''   
    # autoinstaller로 chrome의 webdriver 자동으로 가져오기
    chromedriver_autoinstaller.install()

    # webdriver 실행 (WARNING : chrome 최신버전 상태에서 실행)
    driver = webdriver.Chrome()

    # url로 이동
    url = 'https://www.kdata.or.kr/kr/board/info_01/boardView.do?bbsIdx=33688'
    driver.get(url)

    # 특정 상태를 만족하기까지 기다려 줄 시간 설정
    wait = WebDriverWait(driver, 5)

    # excel 다운로드 링크가 클릭 가능할때까지 대기
    download = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[@class='down_box']/p")))

    # excel 다운로드 링크 클릭
    download.click()

    # 다운로드가 될 때까지 기다림. time.sleep으로 구체적인 시간을 주는 것 말고도 다른 방법도 있을 수 있음
    time.sleep(10) 


def get_excel2():
    '''requests로 웹에서 excel 가져오기 (파일저장X)'''
    # 웹에 있는 엑셀파일 주소
    url = 'https://www.kdata.or.kr/fileDownload.do?srvFile=20230418173119423601.xlsx&usrFile=2022%eb%85%84_%eb%8d%b0%ec%9d%b4%ed%84%b0%ec%82%b0%ec%97%85%ec%a1%b0%ec%82%ac_%ed%86%b5%ea%b3%84%ec%b0%b8%ea%b3%a0%ec%9a%a9table(230418).xlsx&folder=info_01'
    
    # request로 주소에 있는 내용을 가져옴
    response = requests.get(url)

    # BytesIO를 통해 requests로 가져온 내용을 bytes 형식으로 변환
    content = BytesIO(response.content)

    # bytes형식의 엑셀 파일을 openpyxl을 이용하여 읽음
    wb = load_workbook(content, data_only=True)

    # 읽어낸 파일 리턴
    return wb


def excel_to_df(excel: Workbook) -> List:
    '''
    excel을 DataFrame으로 만들기.
    기본적인 흐름이 이런 형식이고,
    실제로 사용하기에는 현재 엑셀 파일과 맞지 않음.
    '''
    df_list = []
    for sheet_name in excel.sheetnames:
        sheet = excel[sheet_name]                   # excel sheet 이름을 통해 sheet에 접근
        data = sheet.values                         # data에 sheet 값을 저장
        columns = next(data)[1:]                    # 첫 번째 열을 헤더로 사용
        df = pd.DataFrame(data, columns=columns)    # DataFrame 생성
        df_list.append(df)                          # 리스트에 생성한 DataFrame 추가

    return df_list                                  # sheet들의 내용을 리스트 형태로 리턴

def excel_to_df2(excel: Workbook) -> List:
    '''excel을 DataFrame으로 만들기'''
    # 표 크기를 측정할 때 병합된 셀이 None값으로 읽히는 문제에 대한 해결
    # 병합된 셀을 모두 해제하는 것으로 해결
    excel = split_cells(excel)
    
    df_list = []
    for sheet_name in excel.sheetnames:
        sheet = excel[sheet_name]                   # excel sheet 이름을 통해 sheet에 접근
        sheet["A3"]
        if sheet_name == '기업수':
            value = table_value(sheet, 'A3')
            df_list.append(value)
        elif sheet_name == '시장규모':
            pass
        elif sheet_name == '인력':
            pass
        else:
            raise ValueError("값이 맞지 않습니다")

    return df_list
    
def table_value(sheet, start_cell):
    '''테이블 위치 기반으로 데이터 가져오는 함수'''
    str_col, start_row = utils.cell.coordinate_from_string("D3")
    start_col = utils.cell.column_index_from_string(str_col)

    # 테이블 사이즈 측정
    num_rows = 0
    num_cols = 0
    row_has_value = False
    col_has_value = False

    for row in sheet.iter_rows(min_row=start_row, min_col=start_col, values_only=True):
        if row_has_value and row[0] is None:
            break
        if row[0] is not None:
            row_has_value = True
        num_rows += 1


    for col in sheet.iter_cols(min_row=start_row, min_col=start_col, values_only=True):
        if col_has_value and col[0] is None:
            break
        if col[0] is not None:
            col_has_value = True
        num_cols += 1
    
    # Read the data from the worksheet
    data = []
    for row in sheet.iter_rows(min_row=start_row, min_col=start_col, max_row=start_row+num_rows-1, max_col=start_col+num_cols-1, values_only=True):
        data.append(row)

    df = pd.DataFrame(data)
    return df

def split_cells(excel: Workbook) -> Workbook:
    '''모든 시트의 병합된 셀을 나누고 값을 채워주는 함수'''
    for sheet_name in excel.sheetnames:
        # 미리 병합된 셀들의 좌표를 저장할 리스트 생성
        merged_cells = []
        sheet = excel[sheet_name]
    
        # 병합된 셀을 확인하고 좌표를 리스트에 저장
        for merged_cell in sheet.merged_cells.ranges:
            merged_cells.append(merged_cell)
    
        # 저장된 병합된 셀들을 순회하면서 분할하여 값을 채우기
        for merged_cell in merged_cells:
            min_col, min_row, max_col, max_row = merged_cell.bounds
            merged_value = sheet.cell(row=min_row, column=min_col).value
    
            # 병합된 셀 해제
            sheet.unmerge_cells(merged_cell.coord)
    
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    sheet.cell(row=row, column=col, value=merged_value)
    
    return excel
    

def df_to_db(df ,db) -> bool:
    '''DataFrame을 DB에 저장하기'''
    

def db_to_df(db, df) -> pd.DataFrame:
    '''DB 데이터를 가져와서 DataFrame 만들기'''
    

def draw():
    '''matplotlib로 그리기'''
    


if __name__ == '__main__':
    #엑셀 가져오기
    excel = get_excel2()
    df_list = excel_to_df2(excel)
    print(df_list[0])
    
    
#%%
# 시트 1(기업수)만 맷플롯립으로 가져오기
df_list[0].plot

#%%
# 0번 째 인덱스를 컬럼으로 변경
df_list[0].rename(columns = df_list[0].loc[0], inplace=True)

#%%
df_list[0].drop([0], axis=0, inplace=True)
# inplace=True < 안하면 원본을 저장하고 진행

#%%
# 인덱스 0번 째 삭제
df_list[0].drop([0], axis=0, inplace=True)

#%%
# 변수명 변경
ndf = df_list[0]
#%%
#14라는 이름의 인덱스 행만 가져오기 (loc)
ndf.loc[14]
#%%


#그래프 뽑기
plt.plot(ndf.loc[14])
# 그래프 제목
plt.title('우리는 할 수 있습니다!')
# X축 이름
plt.xlabel('년도별')
# Y축 이름
plt.ylabel('년도별 기업의 직원 채용 수요')
